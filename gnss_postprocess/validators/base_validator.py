# -*- coding: utf-8 -*-
"""
base_validator.py
Valida las coordenadas de base ingresadas por el usuario.
Incluye validación de zona UTM para Perú.
"""
import os
import json
from typing import Tuple, List, Optional, Dict
from ..gnss_engine.coord_converter import CoordConverter, BaseCoords


class BaseCoordValidator:
    """
    Valida y construye el objeto BaseCoords desde cualquier formato de entrada.
    Soporta: formulario manual, CSV, Excel, JSON.
    """

    ZONAS_PERU = {'17S', '18S', '19S', '18N', '19N'}  # Zonas válidas para territorio peruano

    def __init__(self):
        self._converter = CoordConverter()

    # ══════════════════════════════════════════════
    # DESDE FORMULARIO MANUAL (UTM)
    # ══════════════════════════════════════════════
    def from_utm_form(self,
                      este: float, norte: float, zona: str,
                      h_elip: float, datum: str = 'WGS84',
                      rinex_lat: float = None, rinex_lon: float = None,
                      rinex_h: float = None) -> Tuple[Optional[BaseCoords], List[str]]:
        errors = []

        # Validar zona
        zona_up = zona.strip().upper()
        if zona_up not in self.ZONAS_PERU:
            errors.append(
                f'Zona UTM {zona_up!r} no reconocida para Perú. '
                f'Válidas: {sorted(self.ZONAS_PERU)}'
            )
            return None, errors

        # Validar rangos UTM aproximados para Perú
        if not (200_000 <= este <= 900_000):
            errors.append(f'Este UTM fuera de rango: {este:.3f} m (esperado 200000-900000)')
        if not (7_500_000 <= norte <= 10_200_000):
            errors.append(f'Norte UTM fuera de rango: {norte:.3f} m (esperado 7500000-10200000)')

        if errors:
            return None, errors

        # Convertir a geográficas
        try:
            lat, lon, h = self._converter.utm_to_geo(este, norte, zona_up, h_elip)
        except Exception as ex:
            errors.append(f'Error en conversión UTM→Geo: {ex}')
            return None, errors

        bc = BaseCoords(
            lat_dd=lat, lon_dd=lon, h_elip=h_elip,
            datum=datum, fuente='formulario_utm',
            zona_utm=zona_up, este_utm=este, norte_utm=norte,
            rinex_lat=rinex_lat, rinex_lon=rinex_lon, rinex_h=rinex_h
        )
        return bc, []

    # ══════════════════════════════════════════════
    # DESDE FORMULARIO MANUAL (GEOGRÁFICAS DMS)
    # ══════════════════════════════════════════════
    def from_geo_dms_form(self,
                          lat_d: int, lat_m: int, lat_s: float, lat_hem: str,
                          lon_d: int, lon_m: int, lon_s: float, lon_hem: str,
                          h_elip: float, datum: str = 'WGS84',
                          rinex_lat=None, rinex_lon=None, rinex_h=None
                          ) -> Tuple[Optional[BaseCoords], List[str]]:
        errors = []
        lat_dd = CoordConverter.dms_to_dd(lat_d, lat_m, lat_s, lat_hem)
        lon_dd = CoordConverter.dms_to_dd(lon_d, lon_m, lon_s, lon_hem)

        e, _ = self._validate_geo_range(lat_dd, lon_dd)
        errors.extend(e)
        if errors:
            return None, errors

        bc = BaseCoords(
            lat_dd=lat_dd, lon_dd=lon_dd, h_elip=h_elip,
            datum=datum, fuente='formulario_dms',
            rinex_lat=rinex_lat, rinex_lon=rinex_lon, rinex_h=rinex_h
        )
        return bc, []

    # ══════════════════════════════════════════════
    # DESDE GEOGRÁFICAS DECIMALES
    # ══════════════════════════════════════════════
    def from_geo_decimal(self,
                         lat_dd: float, lon_dd: float, h_elip: float,
                         datum: str = 'WGS84',
                         rinex_lat=None, rinex_lon=None, rinex_h=None
                         ) -> Tuple[Optional[BaseCoords], List[str]]:
        errors, _ = self._validate_geo_range(lat_dd, lon_dd)
        if errors:
            return None, errors
        bc = BaseCoords(
            lat_dd=lat_dd, lon_dd=lon_dd, h_elip=h_elip,
            datum=datum, fuente='formulario_decimal',
            rinex_lat=rinex_lat, rinex_lon=rinex_lon, rinex_h=rinex_h
        )
        return bc, []

    # ══════════════════════════════════════════════
    # DESDE CARTESIANAS ECEF
    # ══════════════════════════════════════════════
    def from_ecef(self, X: float, Y: float, Z: float,
                  datum: str = 'WGS84',
                  rinex_lat=None, rinex_lon=None, rinex_h=None
                  ) -> Tuple[Optional[BaseCoords], List[str]]:
        errors = []
        try:
            lat, lon, h = self._converter.ecef_to_geo(X, Y, Z)
        except Exception as ex:
            errors.append(f'Error ECEF→Geo: {ex}')
            return None, errors
        e, _ = self._validate_geo_range(lat, lon)
        errors.extend(e)
        if errors:
            return None, errors
        bc = BaseCoords(
            lat_dd=lat, lon_dd=lon, h_elip=h,
            datum=datum, fuente='ecef',
            rinex_lat=rinex_lat, rinex_lon=rinex_lon, rinex_h=rinex_h
        )
        return bc, []

    # ══════════════════════════════════════════════
    # DESDE ARCHIVO CSV/JSON/EXCEL
    # ══════════════════════════════════════════════
    def from_file(self, path: str,
                  rinex_lat=None, rinex_lon=None, rinex_h=None
                  ) -> Tuple[Optional[BaseCoords], List[str]]:
        errors = []
        if not os.path.isfile(path):
            return None, [f'Archivo no encontrado: {path}']

        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == '.json':
                return self._from_json(path, rinex_lat, rinex_lon, rinex_h)
            elif ext == '.csv':
                return self._from_csv(path, rinex_lat, rinex_lon, rinex_h)
            elif ext in ('.xlsx', '.xls'):
                return self._from_excel(path, rinex_lat, rinex_lon, rinex_h)
            else:
                return None, [f'Formato no soportado: {ext}. Usa .csv, .json o .xlsx']
        except Exception as ex:
            return None, [f'Error leyendo archivo: {ex}']

    def _from_json(self, path, rl, rlo, rh):
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return self._from_dict(data, rl, rlo, rh)

    def _from_csv(self, path, rl, rlo, rh):
        import csv
        with open(path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        if not rows:
            return None, ['CSV vacío']
        return self._from_dict({k.strip().lower(): v for k, v in rows[0].items()}, rl, rlo, rh)

    def _from_excel(self, path, rl, rlo, rh):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            headers = [str(c.value).strip().lower() for c in ws[1]]
            row = {headers[i]: ws[2][i].value for i in range(len(headers))}
            return self._from_dict(row, rl, rlo, rh)
        except ImportError:
            return None, ['openpyxl no instalado. pip install openpyxl']

    def _from_dict(self, d: Dict, rl, rlo, rh):
        """
        Campos aceptados en el dict (case-insensitive):
          lat/latitude, lon/longitude, h/height/altura,
          este/easting, norte/northing, zona/zone,
          x, y, z (ECEF)
        """
        errors = []
        d = {k.lower(): v for k, v in d.items()}

        # Intento UTM
        if 'este' in d or 'easting' in d:
            este  = float(d.get('este') or d.get('easting'))
            norte = float(d.get('norte') or d.get('northing'))
            zona  = str(d.get('zona') or d.get('zone', '18S'))
            h     = float(d.get('h') or d.get('height') or d.get('altura', 0))
            return self.from_utm_form(este, norte, zona, h,
                                      rinex_lat=rl, rinex_lon=rlo, rinex_h=rh)

        # Intento Geográficas decimales
        if 'lat' in d or 'latitude' in d:
            lat = float(d.get('lat') or d.get('latitude'))
            lon = float(d.get('lon') or d.get('longitude'))
            h   = float(d.get('h') or d.get('height') or d.get('altura', 0))
            return self.from_geo_decimal(lat, lon, h,
                                         rinex_lat=rl, rinex_lon=rlo, rinex_h=rh)

        # Intento ECEF
        if 'x' in d:
            X, Y, Z = float(d['x']), float(d['y']), float(d['z'])
            return self.from_ecef(X, Y, Z, rinex_lat=rl, rinex_lon=rlo, rinex_h=rh)

        return None, ['No se reconocieron campos de coordenadas en el archivo.']

    # ──────────────────────────────────────────────
    # HELPERS
    # ──────────────────────────────────────────────
    @staticmethod
    def _validate_geo_range(lat: float, lon: float):
        errors = []
        if not (-20.0 <= lat <= 2.0):
            errors.append(f'Latitud fuera de rango Perú: {lat:.6f}°')
        if not (-82.0 <= lon <= -68.0):
            errors.append(f'Longitud fuera de rango Perú: {lon:.6f}°')
        return errors, None
